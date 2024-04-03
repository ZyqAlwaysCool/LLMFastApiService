'''
Author: zyq
Date: 2024-03-20 14:40:42
LastEditTime: 2024-03-26 09:41:48
FilePath: /giit/Qilin-Med-VL/llava/serve/cli_test.py
Description: 医疗模型推理流水线. 输入一批jpg数据，基于问题输出模型回答，生成excel表
调用命令：python -m llava.serve.cli_test --model-path /home/kemove/zyq/giit/models/Qilin-Med-VL-Chat

Copyright (c) 2024 by zyq, All Rights Reserved. 
'''
import os 

#两卡ok，超过2会报错：RuntimeError: probability tensor contains either `inf`, `nan` or element < 0
os.environ['CUDA_VISIBLE_DEVICES'] = '2,3'

#单卡ok
# os.environ['CUDA_VISIBLE_DEVICES'] = '1'


import argparse
import torch

from llava.constants import IMAGE_TOKEN_INDEX, DEFAULT_IMAGE_TOKEN, DEFAULT_IM_START_TOKEN, DEFAULT_IM_END_TOKEN
from llava.conversation import conv_templates, SeparatorStyle
from llava.model.builder import load_pretrained_model
from llava.utils import disable_torch_init
from llava.mm_utils import process_images, tokenizer_image_token, get_model_name_from_path, KeywordsStoppingCriteria

import PIL
import requests
from PIL import Image
from io import BytesIO
from transformers import TextStreamer

# excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ws_img
from openpyxl.utils import get_column_letter



def get_folders(path:str):
    folders = [folder for folder in os.listdir(path) if os.path.isdir(os.path.join(path, folder))]
    return folders

def get_file(folder_path:str):
    files_info = []
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path):
            file_info = {
                'name': file_name,
                'path': file_path,
            }
            files_info.append(file_info)
    return files_info

def load_image(image_file:str):
    if image_file.startswith('http://') or image_file.startswith('https://'):
        response = requests.get(image_file)
        image = Image.open(BytesIO(response.content)).convert('RGB')
    else:
        image = Image.open(image_file).convert('RGB')
    return image

def gen_excel(folder_to_file_dct:dict, file_to_result_dct:dict):
    res_path = '/home/kemove/zyq/giit/med-data/test_med/'
    col_width = 100
    row_height = 400
    img_size = 500
    img_col = 'A'
    text_col = 'B'
    
    for folder, file_lst in folder_to_file_dct.items():
        wb = Workbook()
        ws = wb.active
        excel_path = res_path + folder + '/' + folder + '_result.xlsx'
        row_cnt = 1
        
        print('start gen excel. folder=({})'.format(folder))
        
        for file in file_lst:
            print('file info=({})'.format(file))
            if file['path'].split('.')[-1] == 'pdf':
                continue
            img_path = file['path']
            img_type = PIL.Image.open(img_path).convert("RGB")
            img_type.save(img_path, 'jpeg')
            img =ws_img(img_path)
            img.width, img.height = (img_size, img_size)
            ws.column_dimensions[img_col].width = col_width
            ws.row_dimensions[row_cnt].height = row_height
            ws.add_image(img, anchor=img_col + str(row_cnt))   
            
            #生成回答文本
            if img_path not in file_to_result_dct.keys():
                print('file path not exist. check it. img_path=({})'.format(img_path))
            real_text = ''
            for i in file_to_result_dct[img_path]:
                ques = i[0]
                anws = i[1]
                r = '提问: {} 回答: {}'.format(ques, anws)
                real_text += r
                real_text += '\n'
            
            ws.column_dimensions[text_col].width = col_width
            ws.cell(row=row_cnt, column=2, value=real_text) #2表示B列
            row_cnt += 1
        
        wb.save(excel_path)
        wb.close()
        print("done gen excel. file path=({})".format(excel_path))
            
def real_do_infer(args, image_file:str, questions:list):
    print('do real infer')
    #1. load Model
    disable_torch_init()
    model_name = 'llava-Qilin-Chat'
    tokenizer, model, image_processor, context_len = load_pretrained_model(args.model_path, args.model_base, model_name, False, False, device=args.device)
        
    conv_mode = 'llava_v1'

    if args.conv_mode is not None and conv_mode != args.conv_mode:
        print('[WARNING] the auto inferred conversation mode is {}, while `--conv-mode` is {}, using {}'.format(conv_mode, args.conv_mode, args.conv_mode))
    else:
        args.conv_mode = conv_mode

    conv = conv_templates[args.conv_mode].copy()
    roles = conv.roles

    image = load_image(image_file)
    # Similar operation in model_worker.py
    image_tensor = process_images([image], image_processor, args)
    if type(image_tensor) is list:
        image_tensor = [image.to(model.device, dtype=torch.float16) for image in image_tensor]
    else:
        image_tensor = image_tensor.to(model.device, dtype=torch.float16)

    # 2. do infere
    cnt = 0 #逐个输入问题列表中的指定问题
    res = []
    while True:
        if cnt == len(questions):
            break
        
        question = questions[cnt]
        inp = question

        if image is not None:
            # first message
            if model.config.mm_use_im_start_end:
                inp = DEFAULT_IM_START_TOKEN + DEFAULT_IMAGE_TOKEN + DEFAULT_IM_END_TOKEN + '\n' + inp
            else:
                inp = DEFAULT_IMAGE_TOKEN + '\n' + inp
            conv.append_message(conv.roles[0], inp)
            image = None
        else:
            # later messages
            conv.append_message(conv.roles[0], inp)
        conv.append_message(conv.roles[1], None)
        prompt = conv.get_prompt()

        input_ids = tokenizer_image_token(prompt, tokenizer, IMAGE_TOKEN_INDEX, return_tensors='pt').unsqueeze(0).cuda()
        stop_str = conv.sep if conv.sep_style != SeparatorStyle.TWO else conv.sep2
        keywords = [stop_str]
        stopping_criteria = KeywordsStoppingCriteria(keywords, tokenizer, input_ids)
        streamer = TextStreamer(tokenizer, skip_prompt=True, skip_special_tokens=True)
        
        with torch.inference_mode():
            output_ids = model.generate(
                input_ids,
                images=image_tensor,
                do_sample=True,
                temperature=args.temperature,
                max_new_tokens=args.max_new_tokens,
                streamer=streamer,
                use_cache=True,
                stopping_criteria=[stopping_criteria])

        outputs = tokenizer.decode(output_ids[0, input_ids.shape[1]:]).strip()
        conv.messages[-1][-1] = outputs

        if args.debug:
            print("\n", {"prompt": prompt, "outputs": outputs}, "\n")
        
        # # 封装结果
        anwser = outputs.split('</s>')[0]
        res.append([question, anwser])
        cnt += 1
    
    # 3. 释放gpu显存
    print('infer done. release gpu')
    del model
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
    
    return res

# 执行推理
def do_inference(args, folder_path:str, questions:list):
    folders = get_folders(folder_path)
    folder_to_file_dct = {} #文件夹到jpg文件映射
    file_to_result_dct = {} #jgp文件到模型结果映射
    
    
    for folder in folders:
        real_path = folder_path + folder
        folder_to_file_dct[folder] = get_file(real_path)
    
    for folder, files_lst in folder_to_file_dct.items():
        print("folder=({})".format(folder))
        for file in files_lst:
            f_path = file['path']
            if (f_path.split('.')[-1] == 'pdf'): #不处理pdf
                print('match pdf file. skip it. file=({})'.format(file))
                continue
            # 仅支持jpg格式
            if (f_path.split('.')[-1] != 'jpg' and f_path.split('.')[-1] != 'JPG'):
                print('file is not jpg. check it. file=({}) folder=({})'.format(file, folder))
                return -1
            
            res = real_do_infer(args, f_path, questions)
            file_to_result_dct[f_path] = res
    
    return folder_to_file_dct, file_to_result_dct

            

def main(args):
    # folder_path = '/home/kemove/zyq/giit/Qilin-Med-VL/playground/figures/'
    folder_path = '/home/kemove/zyq/giit/med-data/test_med/'
    questions = ['这张图片展示了一种什么类型的医学检查？', '在影像检查中检测到了什么?', '根据图中信息，写一段描述性文本。', '根据图中信息，提供合理的诊断和医学建议。']
    
    # 1.模型推理, 生成结果
    folder_to_file_dct, file_to_result_dct = do_inference(args, folder_path, questions) 

    # 2.后处理, 封装数据结果
    gen_excel(folder_to_file_dct, file_to_result_dct)
    
    
    

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--model-path", type=str, default="facebook/opt-350m")
    parser.add_argument("--model-base", type=str, default=None)
    parser.add_argument("--device", type=str, default="cuda")
    parser.add_argument("--conv-mode", type=str, default=None)
    parser.add_argument("--temperature", type=float, default=0.2)
    parser.add_argument("--max-new-tokens", type=int, default=512)
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--image-aspect-ratio", type=str, default='pad')
    args = parser.parse_args()
    main(args)
